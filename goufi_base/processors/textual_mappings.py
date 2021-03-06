# -*- coding: utf-8 -*-
'''
Created on 23 deb. 2018

@author: C. Guychard
@copyright: ©2018 Article 714
@license: AGPL v3
'''

from openpyxl.cell.read_only import EmptyCell
from openpyxl.reader.excel import load_workbook
from os import path
import re
import unicodecsv
import xlrd

from odoo import _

from odoo.addons.goufi_base.utils.converters  import toString

from .processor import AbstractProcessor

#-------------------------------------------------------------------------------------
# CONSTANTS
XL_AUTHORIZED_EXTS = ('xlsx', 'xls')
CSV_AUTHORIZED_EXTS = ('csv')

#-------------------------------------------------------------------------------------
# MAIN CLASS


class Processor(AbstractProcessor):
    """
    TODO: translate documentation
    TODO: optimize perfs by better using cache of header analysis


    Cette classe permet de traiter des fichiers CSV, XLS ou XLSX pour importer les données qu'ils contiennent dans une instance d'Odoo

    Les fichiers sont importés depuis un répertoire (y compris les sous-répertoires) après avoir été trié par ordre alphabétique

    POUR LES FICHIERS XLS/XLSX

        * chaque feuille est importée séparément
        * le nom du modèle cible est le nom de la feuille

    POUR LES FICHIERS CSV

        * le nom du fichier est du format XX_NOM.csv
             -- XX contenant 2 chiffres pour ordonner l'import,
             -- NOM contient le nom du modèle cible en remplaçant les '.' par des '_'

             /ex, pour importer des res.partner le fichier sera nommé : 00_res_partner.csv

    TRAITEMENT DES LIGNES d'ENTETES (1ere ligne de la feuille (xls*) ou du fichier (csv)

        * la première ligne du fichier définit la façon dont la valeur des colonnes seront importées/traitées:

            -- si la cellule contient le nom d'un champs du modèle, sans modificateur, alors les données de la colonne sont affectés au champs du même nom
                    dans la base
                    /ex. si le modèle cible est res.partner, le contenu de la colonne 'name', sera affecté au champ name de l'enregistrement correspondant

            -- toutes les colonnes contenant des valeurs non retrouvées dans le modèle sont ignorées

            -- Le contenu de la colonne peut démarrer par un modificateur

                    => ">nom_champs/nom_modele_lie/nom_champs_recherche&(filtre)"
                            le champs a adresser (nom_champs) est un champs relationnel vers le modèle (nom_model_lie),
                            de type Many2One
                            pour retrouver l'enregistrement lié, on construit une chaine de recherche avec
                            [noms_champs_recherche=valeur_colonne], le filtre étant utilisé pour restreindre encore la recherche

                            si un enregistrement est trouvé dans le modèle lié, alors l'enregistrement courant est mis à jour/créé avec
                            la valeur "nom_champs" pointant vers ce dernier, sinon, une erreur est remontée et la valeur du champs est ignorée

                    => "*nom_champs/nom_modele_lie/nom_champs_recherche&(filtre)"
                            le champs a adresser (nom_champs) est un champs relationnel vers le modèle (nom_model_lie),
                            de type One2Many
                            pour retrouver les enregistrements liés, on itère sur toutes les valeurs (séparées par des ';' points virgules)
                            pour construire une chaine de recherche avec [noms_champs_recherche=valeur_colonne],
                            le filtre étant utilisé pour restreindre encore la recherche

                            pour chaque enregistrement trouvé dans le modèle lié, l'enregistrement courant est mis à jour/créé en ajoutant
                            à la collection  "nom_champs" un pointeur vers l'enregistrement cible

                    => "+nom_champs/nom_modele_lie"

                            le champs a adresser (nom_champs) est un champs relationnel vers le modèle (nom_model_lie),
                            de type Many2One

                            on crée un ou plusieurs enregistrements cibles avec les valeurs données sous forme de dictionnaire
                            (/ex.: {'name':'Mardi','dayofweek':'1',...})
                            séparés par des ';' points virgules

    """

    def __init__(self, parent_config):

        super(Processor, self).__init__(parent_config)

        # variables use during processing
        self.o2mFields = {}
        self.m2oFields = {}
        self.stdFields = []
        self.idFields = []
        self.allFields = {}
        self.delOrArchFields = {}

        self.header_line_idx = self.parent_config.default_header_line_index
        self.target_model = None

    #-------------------------------------------------------------------------------------
    # process a line of data

    def map_values(self, row):
        for f in row.keys():
            if f in self.allFields:
                # replace non json-compatible values
                val = row[f]
                if val == "False" or val == "True":
                    val = eval(val)
                elif val == None:
                    del(row[f])
                # replace actual col name by actual field name
                if self.allFields[f] != f:
                    row[self.allFields[f]] = val
                    del(row[f])
                else:
                    row[f] = val
            else:
                del(row[f])
        return row

    #-------------------------------------------------------------------------------------
    # Process mappings configuration for each tab

    def process_header(self, header_lines = [], tab_name = None):

        self.o2mFields = {}
        self.m2oFields = {}
        self.stdFields = []
        self.idFields = {}
        self.allFields = {}
        self.delOrArchFields = {}
        col_mappings = None

        tabmap_model = self.odooenv['goufi.tab_mapping']

        # Look for target Model in parent config
        if self.parent_config.tab_support:
            if tab_name != None:
                found = tabmap_model.search([('parent_configuration', '=', self.parent_config.id), ('name', '=', tab_name)], limit = 1)
                if len(found) == 1:
                    try:
                        self.target_model = self.odooenv[found[0].target_object.model]
                        col_mappings = found[0].column_mappings
                        self.header_line_idx = found[0].default_header_line_index
                    except:
                        self.logger.exception("Target model not found for " + toString(tab_name))
                        return -1
                else:
                    self.logger.error("Tab not found: " + toString(tab_name))
                    return -1
            else:
                self.logger.error("No tab name given")
                return -1
        else:
            col_mappings = self.parent_config.column_mappings
            if self.target_model == None:
                try:
                    self.target_model = self.odooenv[self.parent_config.target_object.model]
                except:
                    self.logger.exception("Tab not found: " + toString(tab_name))
                    return -1

        # We should have a model now
        if self.target_model == None:
            self.logger.error("MODEL NOT FOUND ")
            return -1
        else:
            self.logger.info("NEW SHEET:  Import data for model " + toString(self.target_model._name))

        # List of fields in target model
        target_fields = None
        if self.target_model == None:
            raise Exception('FAILED', "FAILED => NO TARGET MODEL FOUND")
        else:
            target_fields = self.target_model.fields_get_keys()

        # process column mappings
        for val in col_mappings:

            if val.is_identifier:
                if val.mapping_expression in target_fields:
                    self.idFields[val.name] = val.mapping_expression
                    self.stdFields.append(val.name)
                    self.allFields[val.name] = val.mapping_expression
                else:
                    self.logger.debug(toString(val.mapping_expression) + "  -> field not found, IGNORED")
            elif val.is_deletion_marker:
                self.delOrArchFields[val.name] = (True, val.delete_if_expression, val.archive_if_not_deleted)
            elif val.is_archival_marker:
                self.delOrArchFields[val.name] = (False, val.archive_if_expression)
            elif re.match(r'\*.*', val.mapping_expression):
                v = val.mapping_expression.replace('*', '')
                vals = [0] + v.split('/')
                v = vals[1]
                if v in target_fields:
                    self.o2mFields[val.name] = vals
                    self.allFields[val.name] = v
                else:
                    self.logger.debug(toString(v) + "  -> field not found, IGNORED")
            elif re.match(r'\+.*', val.mapping_expression):
                v = val.mapping_expression.replace('+', '')
                vals = [1] + v.split('/')
                v = vals[1]
                if v in target_fields:
                    self.o2mFields[val.name] = vals
                    self.allFields[val.name] = v
                else:
                    self.logger.debug(toString(v) + "  -> field not found, IGNORED")
            elif re.match(r'\>.*', val.mapping_expression):
                v = val.mapping_expression.replace('>', '')
                vals = [0] + v.split('/')
                v = vals[1]
                if v in target_fields:
                    self.m2oFields[val.name] = vals
                    self.allFields[val.name] = v
                    if re.match(r'.*\&.*', self.m2oFields[val.name][3]):
                        (_fieldname, cond) = self.m2oFields[val.name][3].split('&')
                        self.m2oFields[val.name][3] = _fieldname
                        self.m2oFields[val.name].append(eval(cond))
                else:
                    self.logger.debug(toString(v) + "  -> field not found, IGNORED")
            else:
                if val.mapping_expression in target_fields:
                    self.stdFields.append(val.name)
                    self.allFields[val.name] = val.mapping_expression
                else:
                    self.logger.debug(toString(val.mapping_expression) + "  -> field not found, IGNORED")
        return len(self.stdFields) + len(self.idFields) + len(self.m2oFields) + len(self.o2mFields)

    #-------------------------------------------------------------------------------------
    # Process line values
    def process_values(self, filename, line_index, data_values):

        DEFAULT_LOG_STRING = "<" + toString(filename) + "> [ line " + toString(line_index + 1) + "] -> "

        currentObj = None
        TO_BE_ARCHIVED = False
        TO_BE_DELETED = False

        if self.target_model == None:
            return False

        # If there exists an id field we can process deletion, archival and updates
        # if there is no, we can only process creation

        if len(self.idFields) > 0 and self.target_model != None:

            # Detects if record needs to be deleted or archived
            CAN_BE_ARCHIVED = ('active' in self.target_model.fields_get_keys())
            for f in self.delOrArchFields:
                if f in data_values:
                    config = self.delOrArchFields[f]
                    if config[0]:
                        # deletion field
                        TO_BE_DELETED = (re.match(config[1], data_values[f]) != None)
                        TO_BE_ARCHIVED = TO_BE_DELETED and config[2]
                        if TO_BE_ARCHIVED and not CAN_BE_ARCHIVED:
                            self.logger.error("This kind of records can not be archived")
                            TO_BE_ARCHIVED = False
                    else :
                        # archival field
                        TO_BE_ARCHIVED = (re.match(config[1], data_values[f]) != None)
                        if TO_BE_ARCHIVED and not CAN_BE_ARCHIVED:
                            self.logger.error("This kind of records can not be archived")
                            TO_BE_ARCHIVED = False

            # calcul des critères de recherche
            search_criteria = []

            for k in self.idFields:
                keyfield = self.idFields[k]
                if k in data_values:
                    value = data_values[k]
                else:
                    value = None

                if value != None and value != str(''):
                    search_criteria.append((keyfield, '=', value))
                else:
                    self.logger.warning("GOUFI: Do not process line n.%d, as Id column is empty" % (line_index + 1,))
                    return

            # ajout d'une clause pour rechercher tous les enregistrements
            if CAN_BE_ARCHIVED:
                search_criteria.append('|')
                search_criteria.append(('active', '=', True))
                search_criteria.append(('active', '=', False))

            # recherche d'un enregistrement existant
            found = self.target_model.search(search_criteria)

            if len(found) == 1:
                currentObj = found[0]
            elif len(found) > 1:
                self.logger.warning(DEFAULT_LOG_STRING + "FOUND TOO MANY RESULT FOR " + toString(self.target_model) + " with " + toString(search_criteria) + "=>   [" + toString(len(found)) + "]")
                return
            else:
                currentObj = None

        # hook for objects needing to be marked as processed
        # by import
        if not currentObj == None and ('import_processed' in self.target_model.fields_get_keys()):
            currentObj.write({'import_processed':True})
            currentObj.import_processed = True
            self.odooenv.cr.commit()

        # processing archives or deletion and returns
        if TO_BE_DELETED:
            if not currentObj == None:
                try:
                    currentObj.unlink()
                except:
                    if TO_BE_ARCHIVED:
                        self.odooenv.cr.rollback()
                        self.logger.warning("Archiving record as it can not be deleted (line n. %d)" % (line_index + 1,))
                        try:
                            currentObj.write({'active':False})
                            currentObj.active = False
                        except Exception as e:
                            self.odooenv.cr.rollback()
                            self.logger.warning("Not able to archive record (line n. %d) : %s" % (line_index + 1, toString(e),))
                currentObj = None
                self.odooenv.cr.commit()
            return
        elif TO_BE_ARCHIVED:
            if not currentObj == None:
                try:
                    currentObj.write({'active':False})
                    currentObj.active = False
                    self.odooenv.cr.commit()
                except Exception as e:
                    self.odooenv.cr.rollback()
                    self.logger.warning("Not able to archive record (line n. %d) : %s" % (line_index + 1, toString(e),))
            return

        # Processing of relational fields
        if len(self.o2mFields) > 0 or len(self.m2oFields) > 0:

            stdRow = {}
            for f in self.stdFields:
                if f in data_values:
                    stdRow[f] = data_values[f]

            # Many To One Fields, might be mandatory, so needs to be treated first and added to StdRow
            for f in self.m2oFields.keys():

                if f in data_values:
                    field = self.m2oFields[f]
                    # reference Many2One,
                    if field[0] == 0 and data_values[f] and len(data_values[f]) > 0:
                        cond = []
                        if len(field) > 4:
                            cond = ['&', (field[3], '=', data_values[f]), field[4]]
                        else:
                            cond = [(field[3], '=', data_values[f])]

                        vals = self.odooenv[field[2]].search(cond, limit = 1)

                        if len(vals) == 1:
                            stdRow[f] = vals[0].id
                        else:
                            self.logger.warning(DEFAULT_LOG_STRING + " found " + toString(len(vals)) + " values for " + toString(data_values[f]) + "  unable to reference " + toString(field[3]) + " " + toString(vals))

            # Create Object if it does not yet exist, else, write updates
            try:
                if currentObj == None:
                    currentObj = self.target_model.create(self.map_values(stdRow))
                else:
                    currentObj.write(self.map_values(stdRow))

                self.odooenv.cr.commit()
            except ValueError as e:
                self.odooenv.cr.rollback()
                self.logger.exception(DEFAULT_LOG_STRING + " wrong values where creating/updating object: " + self.target_model.name + " -> " + toString(stdRow) + "[" + toString(currentObj) + "]")
                self.logger.error("                    MSG: {0}".format(toString(e)))
                currentObj = None
            except Exception as e:
                self.odooenv.cr.rollback()
                self.logger.exception(DEFAULT_LOG_STRING + " Generic Error raised Exception")
                currentObj = None

            # One2Many Fields,

            try:
                for f in self.o2mFields.keys():
                    if f in data_values:
                        members = data_values[f].split(';')
                        field = self.o2mFields[f]
                        if len(members) > 0 and currentObj != None:
                            if field[0] == 1:
                                currentObj.write({field[1]:[(5, False, False)] })
                            for m in members:
                                if len(m) > 0:
                                    # For adding some stuffs in lists do : https://www.odoo.com/documentation/10.0/reference/orm.html#odoo.models.Model.write
                                    # References records in  One2Many
                                    if field[0] == 0:
                                        vals = self.odooenv[field[2]].search([(field[3], '=', m)], limit = 1)
                                        if len(vals) == 1:
                                            currentObj.write({field[1]:[(4, vals[0].id, False)] })
                                        else:
                                            self.logger.warning(DEFAULT_LOG_STRING + "found " + toString(len(vals)) + " values for " + toString(m) + "  unable to reference")

                                    # Creates records in  One2Many
                                    elif field[0] == 1:
                                        values = eval(m)
                                        currentObj.write({field[1]:[(0, False, values)] })
                self.odooenv.cr.commit()
            except ValueError as e:
                self.odooenv.cr.rollback()
                self.logger.exception(DEFAULT_LOG_STRING + " Wrong values where updating object: " + self.target_model.name + " -> " + toString(stdRow))
                self.logger.error("                    MSG: {0}".format(toString(e)))
                currentObj = None
            except Exception as e:
                self.odooenv.cr.rollback()
                self.logger.exception(DEFAULT_LOG_STRING + " Generic Error raised Exception")
                currentObj = None

            self.odooenv.cr.commit()

        # pas de champs collection
        else:

            try:
                if self.target_model != None:
                    if currentObj == None :
                        currentObj = self.target_model.create(self.map_values(data_values))
                    else:
                        currentObj.write(self.map_values(data_values))
                    self.odooenv.cr.commit()
                else:
                    self.logger.error(DEFAULT_LOG_STRING + " No Target model was identified ")
            except ValueError as e:
                self.odooenv.cr.rollback()
                self.logger.exception(DEFAULT_LOG_STRING + " error where creating/updating object: " + self.target_model.name + " -> " + toString(data_values) + "[" + toString(currentObj) + "]")
                self.logger.error("                    MSG: {0}".format(toString(e)))
            except Exception as e:
                self.odooenv.cr.rollback()
                self.logger.exception(DEFAULT_LOG_STRING + " Generic Error raised Exception")
                currentObj = None

        # *****

        # Finally commit
        self.odooenv.cr.commit()


#-------------------------------------------------------------------------------------
# Process CSV Only
class CSVProcessor(Processor):
    """
    Processes csv files
    """

    #-------------------------------------------------------------------------------------
    def does_file_need_processing(self, import_file):
        """
        Returns true if the given ImportFile is to be processed
        """
        result = super(CSVProcessor, self).does_file_need_processing(import_file)
        ext = import_file.filename.split('.')[-1]
        return ((ext in CSV_AUTHORIZED_EXTS) and result)

    #-------------------------------------------------------------------------------------
    def process_data(self, import_file):
        """
        Method that actually process data
        """
        self.logger.info("PROCESSING CSV FILE :" + import_file.filename)

        processed = False

        # Search for target model

        self.target_model = None
        try:
            self.target_model = self.odooenv[self.parent_config.target_object.model]
        except:
            self.target_model = None

        if self.target_model == None:

            self.logger.warning("No target model set on configuration, attempt to find it from file name")

            bname = path.basename(import_file.filename)
            (modelname, ext) = bname.split('.')

            modelname = modelname.replace('_', '.')
            if re.match(r'[0-9]+\.', modelname):
                modelname = re.sub(r'[0-9]+\.', '', modelname)

            try:
                self.target_model = self.odooenv[modelname]
            except:
                self.target_model = None
                self.logger.exception("Not able to guess target model from filename: " + toString(import_file.filename))
                import_file.processing_result = _(u"Model Not found")
                import_file.processing_status = 'failure'
                self.odooenv.cr.commit()
                return

        # try with , as a delimiter
        with open(import_file.filename, 'rb') as csvfile:
            csv_reader = unicodecsv.DictReader(csvfile, delimiter = ',', quotechar = '\"')
            if (len(csv_reader.fieldnames) > 1):
                if self.process_header(csv_reader.fieldnames) > 0:
                    processed = True
                    idx = 0
                    for row in csv_reader:
                        idx += 1
                        self.process_values(import_file.filename, idx, row)

            csvfile.close()

        # try with ; as a delimiter
        if not processed:
            with open(import_file.filename, 'rb') as csvfile:
                csv_reader = unicodecsv.DictReader(csvfile, delimiter = ';', quotechar = '\"')

                if self.process_header(csv_reader.fieldnames) > 0 :

                    idx = 0
                    for row in csv_reader:
                        idx += 1
                        try:
                            self.process_values(import_file.filename, idx, row)

                        except Exception as e:
                            self.logger.exception(u"Error when processing line N°" + str(idx))
                            import_file.processing_status = 'failure'
                            import_file.processing_result = str(e) + " -- " + e.message
                            self.odooenv.cr.commit()

                csvfile.close()

        self.logger.info("Textual mapping IMPORT; process DATA: " + toString(import_file.filename))


#-------------------------------------------------------------------------------------
# Process XL* Only
class XLProcessor(Processor):
    """
    Processes xls and xlsx files
    """

    #-------------------------------------------------------------------------------------
    def process_xls(self, import_file):
        self.logger.info("PROCESSING XLS FILE :" + import_file.filename)

        wb = xlrd.open_workbook(import_file.filename)
        for sh in wb.sheets():

            self.target_model = None

            # la ligne se sont les intitutlés
            p_ligne = sh.row_values(self.header_line_idx)
            hsize = len(p_ligne)

            if self.process_header(p_ligne, sh.name) > 0 :

                for rownum in range(1, sh.nrows):
                    if rownum > self.header_line_idx:
                        values = {}
                        row_vals = sh.row_values(rownum)
                        for idx in range(0, hsize):
                            values[p_ligne[idx]] = row_vals[idx]
                    try:
                        self.process_values(import_file.filename, rownum, values)
                    except Exception as e:
                        self.logger.exception(u"Error when processing line N°" + str(rownum) + " in " + sh.name)
        return True

    #-------------------------------------------------------------------------------------
    def process_xlsx(self, import_file):
        self.logger.info("PROCESSING XLSX FILE :" + import_file.filename)

        result = True
        wb = load_workbook(import_file.filename, read_only = True, keep_vba = False, guess_types = False, data_only = True)
        for shname in wb.sheetnames:

            sh = wb[shname]
            firstrow = None
            header_values = []
            idx = 0
            nb_fields = 0
            self.target_model = None

            for r in sh:

                if firstrow == None:
                    if idx == self.header_line_idx:
                        firstrow = r
                        for c in firstrow:
                            header_values.append(c.value)
                        nb_fields = self.process_header(header_values, shname)
                        if nb_fields < 1 or self.target_model == None:
                            # do not continue if not able to process headers
                            self.logger.error(u'Unable to process headers for Tab ' + shname)
                            break
                        elif ('import_processed' in self.target_model.fields_get_keys()):
                            # hook for objects needing to be set as processed through import
                            self.odooenv.cr.execute('update ' + toString(self.target_model._table) + ' set import_processed = False')
                            self.odooenv.cr.commit()
                else:
                    if self.target_model == None:
                        break
                    values = {}
                    for c in r:
                        colname = None
                        if not isinstance(c, EmptyCell) and not c.column == None:
                            colname = firstrow[c.column - 1].value
                        if colname != None:
                            values[colname] = c.value
                    try:
                        self.process_values(import_file.filename, idx, values)
                    except Exception as e:
                        self.logger.exception(u"Error when processing line N°" + str(idx) + " in " + shname)

                idx += 1
            if self.target_model != None:
                if  ('import_processed' in self.target_model.fields_get_keys()):
                    # hook for objects needing to be set as processed through import
                    self.odooenv.cr.execute('update ' + toString(self.target_model._table) + ' set import_processed = False')
                    self.odooenv.cr.commit()
            else:
                self.logger.error("Did not process tab " + shname + " correctly")
                result = False
        return result

    #-------------------------------------------------------------------------------------
    def does_file_need_processing(self, import_file):
        """
        Returns true if the given ImportFile is to be processed
        """
        result = super(XLProcessor, self).does_file_need_processing(import_file)
        ext = import_file.filename.split('.')[-1]
        return ((ext in XL_AUTHORIZED_EXTS) and result)

    #-------------------------------------------------------------------------------------
    def process_data(self, import_file):
        """
        Method that actually process data
        """

        self.logger.info("Textual mapping IMPORT; process DATA: " + toString(import_file.filename))
        try:

            if import_file.filename.endswith('.xls'):
                result = self.process_xls(import_file)
            elif import_file.filename.endswith('.xlsx'):
                result = self.process_xlsx(import_file)

            return result

        except Exception as e:
            self.logger.exception("Processing Failed: " + str(e))
            self.odooenv.cr.rollback()
            import_file.processing_status = 'failure'
            import_file.processing_result = str(e)
            self.odooenv.cr.commit()
            return False

